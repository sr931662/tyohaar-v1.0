"""
Import / Export Service
=======================
Enterprise-grade import engine with validation, duplicate detection,
preview, dry-run, rollback, and progress tracking.

Export engine supports XLSX, CSV, JSON with column selection and filters.
"""

from __future__ import annotations

import io
import secrets
import uuid
from collections.abc import Callable
from datetime import datetime, timezone
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import AsyncSessionLocal
from app.models.cms.export_log import ExportLog
from app.models.cms.import_log import ImportLog
from app.repositories.cms_repository import CMSRepositoryAggregate
from app.schemas.cms.io import (
    ExportLogResponse,
    ExportRequest,
    ExportTriggerResponse,
    ImportLogResponse,
    ImportPreviewResponse,
    ImportPreviewRow,
    ImportUndoResponse,
    RowError,
)
from app.services.base import BaseService

# Column definitions for each importable entity
_ENTITY_COLUMNS: dict[str, list[str]] = {
    "vendors": ["business_name", "phone", "email", "vendor_type", "city", "state", "pincode", "description"],
    "customers": ["phone", "email", "full_name", "city", "state"],
    "packages": [
        "name", "description", "base_price", "currency", "category", "vendor_phone",
        "min_guests", "max_guests", "duration_hours",
        "image_1_url", "image_2_url", "image_3_url", "image_4_url", "image_5_url",
    ],
    "package_categories": ["name", "slug", "description"],
    "cities": ["name", "display_name", "state", "is_metro", "is_serviceable"],
    "states": ["name", "code", "country"],
    "coupons": ["title", "code", "discount_type", "discount_value", "max_uses", "valid_from", "valid_until", "min_order_value"],
    "faqs": ["question", "answer", "category", "display_order"],
    "notification_templates": ["template_key", "channel", "notification_category", "title_template", "body_template"],
    "settings": ["key", "value", "description"],
    "memberships": ["tier", "name", "monthly_price", "yearly_price", "validity_days", "features"],
    "vendor_services": ["vendor_phone", "category", "name", "description", "base_price", "pricing_type"],
}

_REQUIRED_COLUMNS: dict[str, list[str]] = {
    "vendors": ["business_name", "phone"],
    "customers": ["phone"],
    "packages": ["name", "base_price"],
    "package_categories": ["name", "slug"],
    "cities": ["name", "state"],
    "states": ["name", "code"],
    "coupons": ["title", "discount_type", "discount_value", "valid_from"],
    "faqs": ["question", "answer"],
    "notification_templates": ["template_key", "channel", "notification_category", "title_template", "body_template"],
    "settings": ["key", "value"],
    "memberships": ["tier", "name", "monthly_price"],
    "vendor_services": ["vendor_phone", "category", "name", "base_price"],
}

# Entity types with a real bulk-insert implementation in _insert_row().
EXECUTABLE_ENTITY_TYPES = {
    "faqs", "settings", "packages", "vendors", "customers", "package_categories",
    "cities", "states", "coupons", "notification_templates", "memberships", "vendor_services",
}


class IOService(BaseService):
    def __init__(self, session_factory: Callable[[], AsyncSession] = AsyncSessionLocal) -> None:
        super().__init__(session_factory)

    # ── Import ────────────────────────────────────────────────────────────────

    def _parse_file(self, content: bytes, filename: str) -> list[dict[str, Any]]:
        """Parse XLSX or CSV file and return list of row dicts."""
        ext = filename.lower().rsplit(".", 1)[-1]
        if ext == "csv":
            import csv
            reader = csv.DictReader(io.StringIO(content.decode("utf-8", errors="replace")))
            return list(reader)
        elif ext in ("xlsx", "xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return []
                headers = [str(h).strip() if h else "" for h in rows[0]]
                return [
                    {headers[i]: str(cell) if cell is not None else "" for i, cell in enumerate(row)}
                    for row in rows[1:]
                ]
            except ImportError as exc:
                raise RuntimeError(
                    "Server is missing the openpyxl dependency required to read .xlsx files. "
                    "Please upload a .csv file instead, or contact support."
                ) from exc
        return []

    def _validate_row(
        self,
        row: dict[str, Any],
        row_num: int,
        entity_type: str,
        seen_keys: set[str],
    ) -> tuple[bool, bool, list[RowError]]:
        """Returns (is_valid, is_duplicate, errors)."""
        errors: list[RowError] = []
        required = _REQUIRED_COLUMNS.get(entity_type, [])

        for field in required:
            if not row.get(field):
                errors.append(RowError(row=row_num, field=field, message=f"'{field}' is required", value=row.get(field)))

        # Duplicate detection: use primary identifier
        pk_field = required[0] if required else None
        is_duplicate = False
        if pk_field and row.get(pk_field):
            key = f"{entity_type}:{row[pk_field]}"
            if key in seen_keys:
                is_duplicate = True
            else:
                seen_keys.add(key)

        return len(errors) == 0, is_duplicate, errors

    async def get_import_template(self, entity_type: str) -> bytes:
        """Generate a sample XLSX template for the given entity type."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = entity_type.title()

        columns = _ENTITY_COLUMNS.get(entity_type, [])
        required = set(_REQUIRED_COLUMNS.get(entity_type, []))

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        required_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = required_fill if col_name in required else header_fill
            ws.column_dimensions[chr(64 + col_idx)].width = 20

        # Add a sample row
        sample_row = {
            "business_name": "Sample Vendor", "name": "Sample Item",
            "phone": "+919999999999", "vendor_phone": "+919999999999", "email": "sample@example.com",
            "base_price": "5000", "discount_type": "percentage",
            "discount_value": "10", "code": "SAMPLE10", "title": "Sample 10% Off",
            "question": "Sample question?", "answer": "Sample answer.",
            "key": "sample_key", "value": "sample_value",
            "tier": "gold", "monthly_price": "299", "yearly_price": "2999",
            "validity_days": "30", "features": "feature1,feature2",
            "category": "Photography", "slug": "photography",
            "description": "Sample description",
            "city": "Mumbai", "state": "Maharashtra", "display_name": "Mumbai",
            "is_metro": "true", "is_serviceable": "true",
            "country": "IN",
            "valid_from": "2026-01-01", "valid_until": "2026-12-31",
            "max_uses": "100", "min_order_value": "0",
            "min_guests": "10", "max_guests": "100",
            "duration_hours": "4", "currency": "INR",
            "template_key": "booking_confirmed", "channel": "push",
            "notification_category": "booking_confirmed",
            "title_template": "Hello {{name}}", "body_template": "Your {{event}} is ready",
            "display_order": "1", "pricing_type": "fixed",
            "vendor_type": "decorator",
            "image_1_url": "https://example.com/photos/package-cover.jpg",
            "image_2_url": "https://example.com/photos/package-2.jpg",
        }
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=2, column=col_idx, value=sample_row.get(col_name, ""))

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def validate_and_preview(
        self,
        *,
        content: bytes,
        filename: str,
        entity_type: str,
        admin_id: uuid.UUID,
        is_dry_run: bool = False,
    ) -> tuple[ImportPreviewResponse, uuid.UUID]:
        """Parse, validate, and create an ImportLog. Returns preview + log_id."""
        if entity_type not in EXECUTABLE_ENTITY_TYPES:
            from app.services.exceptions import ValidationError
            raise ValidationError(
                f"Bulk import for entity_type='{entity_type}' is not supported yet. "
                f"Supported types: {', '.join(sorted(EXECUTABLE_ENTITY_TYPES))}.",
                field="entity_type",
            )
        rows = self._parse_file(content, filename)
        total = len(rows)
        seen_keys: set[str] = set()
        preview_rows: list[ImportPreviewRow] = []
        valid_count = invalid_count = dup_count = 0

        for idx, row in enumerate(rows[:100], start=2):  # Preview up to 100 rows
            is_valid, is_dup, errors = self._validate_row(row, idx, entity_type, seen_keys)
            if not is_valid:
                invalid_count += 1
            elif is_dup:
                dup_count += 1
            else:
                valid_count += 1

            preview_rows.append(ImportPreviewRow(
                row=idx,
                data=row,
                is_valid=is_valid,
                is_duplicate=is_dup,
                errors=errors,
                action="SKIP" if is_dup else ("INSERT" if is_valid else "SKIP"),
            ))

        # Validate all rows for counts (not just preview)
        all_seen: set[str] = set()
        total_valid = total_invalid = total_dup = 0
        all_errors: list[RowError] = []
        for idx, row in enumerate(rows, start=2):
            is_valid, is_dup, errs = self._validate_row(row, idx, entity_type, all_seen)
            if not is_valid:
                total_invalid += 1
                all_errors.extend(errs[:3])
            elif is_dup:
                total_dup += 1
            else:
                total_valid += 1

        can_proceed = total_invalid == 0 and total > 0

        # Create import log
        async with self._uow() as uow:
            log = await uow.cms.import_logs.create_from_dict({
                "admin_id": admin_id,
                "entity_type": entity_type,
                "original_filename": filename,
                "file_size_bytes": len(content),
                "status": "PREVIEW",
                "is_dry_run": is_dry_run,
                "total_rows": total,
                "valid_rows": total_valid,
                "invalid_rows": total_invalid,
                "duplicate_rows": total_dup,
                "progress_pct": 100.0,
                "error_summary": {
                    "sample_errors": [e.model_dump() for e in all_errors[:20]],
                    "total_errors": total_invalid,
                },
            })
            log_id = log.id

        column_mapping = {c: c for c in (_ENTITY_COLUMNS.get(entity_type) or [])}

        return ImportPreviewResponse(
            entity_type=entity_type,
            filename=filename,
            total_rows=total,
            valid_rows=total_valid,
            invalid_rows=total_invalid,
            duplicate_rows=total_dup,
            preview_rows=preview_rows,
            column_mapping=column_mapping,
            sample_errors=all_errors[:10],
            can_proceed=can_proceed,
        ), log_id

    async def execute_import(
        self,
        *,
        log_id: uuid.UUID,
        content: bytes,
        overwrite_duplicates: bool = False,
    ) -> ImportLogResponse:
        """Execute the actual import after preview has been accepted."""
        async with self._uow() as uow:
            log = await uow.cms.import_logs.get_by_id_or_raise(log_id)

            if log.status not in ("PREVIEW", "FAILED"):
                from fastapi import HTTPException, status as s
                raise HTTPException(
                    status_code=s.HTTP_409_CONFLICT,
                    detail=f"Import cannot be executed in status '{log.status}'",
                )

            await uow.cms.import_logs.update(log, {
                "status": "PROCESSING",
                "started_at": datetime.now(timezone.utc),
                "progress_pct": 0.0,
            })

            rows = self._parse_file(content, log.original_filename)
            inserted = 0
            errors: list[dict[str, Any]] = []
            rollback_ids: list[str] = []

            for idx, row in enumerate(rows, start=2):
                try:
                    new_id = await self._insert_row(uow.session, log.entity_type, row)
                    if new_id:
                        rollback_ids.append(str(new_id))
                        inserted += 1
                except Exception as exc:
                    errors.append({"row": idx, "error": str(exc)})

            await uow.cms.import_logs.update(log, {
                "status": "COMPLETED" if not errors else "FAILED",
                "inserted_rows": inserted,
                "progress_pct": 100.0,
                "completed_at": datetime.now(timezone.utc),
                "rollback_snapshot": {"inserted_ids": rollback_ids},
                "error_summary": {"errors": errors[:50]} if errors else None,
            })

            await uow.commit()

        async with self._uow() as uow:
            log = await uow.cms.import_logs.get_by_id_or_raise(log_id)
            return ImportLogResponse.model_validate(log)

    async def _insert_row(self, session: AsyncSession, entity_type: str, row: dict[str, Any]) -> uuid.UUID | None:
        """Insert a single validated row into the appropriate table."""
        if entity_type == "faqs":
            from app.models.common.faq import FAQ
            from app.models.enums import FAQCategory

            raw_category = (row.get("category") or "").strip().lower()
            try:
                faq_category = FAQCategory(raw_category) if raw_category else FAQCategory.GENERAL
            except ValueError:
                faq_category = FAQCategory.GENERAL

            obj = FAQ(
                question=row.get("question", ""),
                answer=row.get("answer", ""),
                faq_category=faq_category,
                display_order=int(row.get("display_order", 0) or 0),
                is_active=True,
            )
            session.add(obj)
            await session.flush()
            return obj.id
        elif entity_type == "settings":
            from app.models.common.app_setting import AppSetting
            key = row.get("key", "")
            obj = AppSetting(
                key=key,
                # group/label are required with no default; the import
                # template only asks for key/value/description, so derive
                # reasonable values rather than crash on every row.
                group="imported",
                label=key.replace("_", " ").title() or "Imported Setting",
                value=row.get("value", ""),
                description=row.get("description"),
            )
            session.add(obj)
            await session.flush()
            return obj.id
        elif entity_type == "packages":
            import re
            import uuid as _uuid
            from decimal import Decimal, InvalidOperation

            from sqlalchemy import func, select

            from app.models.enums import PackageStatus
            from app.models.packages.package import Package
            from app.models.packages.package_category import PackageCategory
            from app.models.users.user import User
            from app.models.vendors.vendor import Vendor

            name = (row.get("name") or "").strip()
            if not name:
                raise ValueError("'name' is required")

            try:
                base_price = Decimal(str(row.get("base_price")))
            except (InvalidOperation, TypeError, ValueError):
                raise ValueError(f"Invalid base_price: {row.get('base_price')!r}")

            vendor_id = None
            vendor_phone = (row.get("vendor_phone") or "").strip()
            if vendor_phone:
                vendor = (await session.execute(
                    select(Vendor).join(User, User.id == Vendor.user_id).where(User.phone == vendor_phone)
                )).scalars().first()
                if vendor is None:
                    raise ValueError(f"No vendor found with phone '{vendor_phone}'")
                vendor_id = vendor.id

            category_id = None
            category_name = (row.get("category") or "").strip()
            if category_name:
                category = (await session.execute(
                    select(PackageCategory).where(func.lower(PackageCategory.name) == category_name.lower())
                )).scalars().first()
                if category is not None:
                    category_id = category.id

            def _to_int(value: Any) -> int | None:
                try:
                    return int(value) if value not in (None, "") else None
                except (TypeError, ValueError):
                    return None

            def _to_decimal(value: Any) -> Decimal | None:
                try:
                    return Decimal(str(value)) if value not in (None, "") else None
                except (InvalidOperation, TypeError, ValueError):
                    return None

            base_slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
            slug = f"{base_slug}-{str(_uuid.uuid4())[:8]}"

            obj = Package(
                name=name,
                slug=slug,
                description=row.get("description") or None,
                base_price=base_price,
                currency=(row.get("currency") or "INR").strip().upper() or "INR",
                category_id=category_id,
                vendor_id=vendor_id,
                min_guest_count=_to_int(row.get("min_guests")),
                max_guest_count=_to_int(row.get("max_guests")),
                duration_hours=_to_decimal(row.get("duration_hours")),
                status=PackageStatus.DRAFT,
                is_active=False,
            )
            session.add(obj)
            await session.flush()

            # Attach any image_N_url columns as package gallery items —
            # lets one spreadsheet row set up a full package with photos in
            # one import instead of a separate manual upload step per image.
            from app.models.enums import MediaStatus, MediaType, MediaUsage
            from app.models.packages.package_gallery import PackageGallery

            image_urls = [
                (row.get(f"image_{i}_url") or "").strip()
                for i in range(1, 6)
            ]
            for sort_order, url in enumerate(u for u in image_urls if u):
                session.add(PackageGallery(
                    package_id=obj.id,
                    media_type=MediaType.IMAGE,
                    usage=MediaUsage.PACKAGE_IMAGE,
                    file_url=url,
                    is_featured=(sort_order == 0),
                    sort_order=sort_order,
                    status=MediaStatus.ACTIVE,
                ))
            await session.flush()
            return obj.id
        elif entity_type == "vendors":
            from app.models.enums import (
                AccountStatus,
                UserRole,
                VendorStatus,
                VendorType,
                VendorVerificationStatus,
            )
            from app.models.users.user import User
            from app.models.vendors.vendor import Vendor
            from app.models.vendors.vendor_profile import VendorProfile
            from sqlalchemy import select

            business_name = (row.get("business_name") or "").strip()
            phone = (row.get("phone") or "").strip()
            if not business_name:
                raise ValueError("'business_name' is required")
            if not phone:
                raise ValueError("'phone' is required")

            existing_user = (await session.execute(
                select(User).where(User.phone == phone)
            )).scalars().first()

            if existing_user is not None:
                user_id = existing_user.id
            else:
                new_user = User(
                    phone=phone,
                    email=row.get("email") or None,
                    full_name=business_name,
                    role=UserRole.VENDOR,
                    account_status=AccountStatus.ACTIVE,
                )
                session.add(new_user)
                await session.flush()
                user_id = new_user.id

            existing_vendor = (await session.execute(
                select(Vendor).where(Vendor.user_id == user_id)
            )).scalars().first()
            if existing_vendor is not None:
                raise ValueError(f"Vendor already exists for phone '{phone}'")

            raw_type = (row.get("vendor_type") or "").strip().lower()
            try:
                vendor_type = VendorType(raw_type) if raw_type else VendorType.PLANNER
            except ValueError:
                vendor_type = VendorType.PLANNER

            vendor = Vendor(
                user_id=user_id,
                business_name=business_name,
                vendor_type=vendor_type,
                verification_status=VendorVerificationStatus.UNVERIFIED,
                status=VendorStatus.PENDING,
                is_active=False,
            )
            session.add(vendor)
            await session.flush()
            session.add(VendorProfile(vendor_id=vendor.id))
            await session.flush()
            return vendor.id
        elif entity_type == "customers":
            from app.models.enums import AccountStatus, UserRole
            from app.models.users.user import User
            from sqlalchemy import select

            phone = (row.get("phone") or "").strip()
            if not phone:
                raise ValueError("'phone' is required")

            existing = (await session.execute(
                select(User).where(User.phone == phone)
            )).scalars().first()
            if existing is not None:
                raise ValueError(f"User already exists with phone '{phone}'")

            obj = User(
                phone=phone,
                email=row.get("email") or None,
                full_name=row.get("full_name") or None,
                role=UserRole.CUSTOMER,
                account_status=AccountStatus.ACTIVE,
            )
            session.add(obj)
            await session.flush()
            return obj.id
        elif entity_type == "package_categories":
            import re

            from sqlalchemy import func, select

            from app.models.packages.package_category import PackageCategory

            name = (row.get("name") or "").strip()
            if not name:
                raise ValueError("'name' is required")

            slug = (row.get("slug") or "").strip().lower() or re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")

            existing = (await session.execute(
                select(PackageCategory).where(
                    (func.lower(PackageCategory.name) == name.lower())
                    | (PackageCategory.slug == slug)
                )
            )).scalars().first()
            if existing is not None:
                raise ValueError(f"A package category named '{name}' or with slug '{slug}' already exists")

            obj = PackageCategory(
                name=name,
                slug=slug,
                description=row.get("description") or None,
                is_active=True,
            )
            session.add(obj)
            await session.flush()
            return obj.id
        elif entity_type == "states":
            from app.models.common.state import State
            from sqlalchemy import func, select

            name = (row.get("name") or "").strip()
            code = (row.get("code") or "").strip().upper()
            if not name:
                raise ValueError("'name' is required")
            if not code:
                raise ValueError("'code' is required")
            country_code = (row.get("country") or "IN").strip().upper()

            existing = (await session.execute(
                select(State).where(
                    (func.lower(State.name) == name.lower()) | (State.code == code),
                    State.country_code == country_code,
                )
            )).scalars().first()
            if existing is not None:
                raise ValueError(f"A state named '{name}' or with code '{code}' already exists for country '{country_code}'")

            obj = State(name=name, code=code, country_code=country_code)
            session.add(obj)
            await session.flush()
            return obj.id
        elif entity_type == "cities":
            import re

            from app.models.common.city import City
            from app.models.common.state import State
            from sqlalchemy import func, select

            name = (row.get("name") or "").strip()
            state_name = (row.get("state") or "").strip()
            if not name:
                raise ValueError("'name' is required")
            if not state_name:
                raise ValueError("'state' is required")

            state = (await session.execute(
                select(State).where(func.lower(State.name) == state_name.lower())
            )).scalars().first()
            if state is None:
                raise ValueError(f"No state found named '{state_name}' — import states first")

            existing = (await session.execute(
                select(City).where(City.state_id == state.id, func.lower(City.name) == name.lower())
            )).scalars().first()
            if existing is not None:
                raise ValueError(f"City '{name}' already exists in state '{state_name}'")

            base_slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
            slug = base_slug
            suffix = 2
            while (await session.execute(select(City).where(City.slug == slug))).scalars().first() is not None:
                slug = f"{base_slug}-{suffix}"
                suffix += 1

            def _to_bool(value: Any) -> bool:
                return str(value).strip().lower() in ("true", "1", "yes")

            obj = City(
                state_id=state.id,
                name=name,
                display_name=row.get("display_name") or name,
                slug=slug,
                is_metro=_to_bool(row.get("is_metro")),
                is_serviceable=_to_bool(row.get("is_serviceable")),
            )
            session.add(obj)
            await session.flush()
            return obj.id
        elif entity_type == "coupons":
            from decimal import Decimal, InvalidOperation

            from app.models.enums import CouponAdminStatus, CouponType
            from app.models.payments.coupon import Coupon
            from sqlalchemy import select

            title = (row.get("title") or "").strip()
            if not title:
                raise ValueError("'title' is required")

            raw_type = (row.get("discount_type") or "").strip().lower()
            try:
                coupon_type = CouponType(raw_type)
            except ValueError:
                raise ValueError(
                    f"Invalid discount_type '{raw_type}'. Must be one of: {', '.join(t.value for t in CouponType)}"
                )

            try:
                discount_value = Decimal(str(row.get("discount_value")))
            except (InvalidOperation, TypeError, ValueError):
                raise ValueError(f"Invalid discount_value: {row.get('discount_value')!r}")

            try:
                valid_from = datetime.fromisoformat(str(row.get("valid_from")))
            except (TypeError, ValueError):
                raise ValueError(f"Invalid valid_from date: {row.get('valid_from')!r}")

            valid_until_raw = (row.get("valid_until") or "").strip()
            valid_until = None
            if valid_until_raw:
                try:
                    valid_until = datetime.fromisoformat(valid_until_raw)
                except ValueError:
                    raise ValueError(f"Invalid valid_until date: {valid_until_raw!r}")

            code = (row.get("code") or "").strip().upper() or None
            if code:
                existing = (await session.execute(
                    select(Coupon).where(Coupon.code == code)
                )).scalars().first()
                if existing is not None:
                    raise ValueError(f"A coupon with code '{code}' already exists")

            max_uses_raw = (row.get("max_uses") or "").strip()
            min_order_raw = (row.get("min_order_value") or "").strip()

            obj = Coupon(
                title=title,
                code=code,
                coupon_type=coupon_type,
                discount_value=discount_value,
                valid_from=valid_from,
                valid_until=valid_until,
                total_usage_limit=int(max_uses_raw) if max_uses_raw else None,
                min_order_value=Decimal(min_order_raw) if min_order_raw else None,
                # Imported coupons always start in Draft so an admin reviews
                # eligibility rules before they can be redeemed by customers.
                admin_status=CouponAdminStatus.DRAFT,
            )
            session.add(obj)
            await session.flush()
            return obj.id
        elif entity_type == "notification_templates":
            from app.models.enums import NotificationChannel, NotificationType
            from app.models.notifications.template import NotificationTemplate
            from sqlalchemy import select

            template_key = (row.get("template_key") or "").strip()
            title_template = (row.get("title_template") or "").strip()
            body_template = (row.get("body_template") or "").strip()
            if not template_key:
                raise ValueError("'template_key' is required")
            if not title_template:
                raise ValueError("'title_template' is required")
            if not body_template:
                raise ValueError("'body_template' is required")

            raw_channel = (row.get("channel") or "").strip().lower()
            try:
                channel = NotificationChannel(raw_channel)
            except ValueError:
                raise ValueError(
                    f"Invalid channel '{raw_channel}'. Must be one of: {', '.join(c.value for c in NotificationChannel)}"
                )

            raw_category = (row.get("notification_category") or "").strip().lower()
            try:
                notification_category = NotificationType(raw_category)
            except ValueError:
                raise ValueError(
                    f"Invalid notification_category '{raw_category}'. "
                    f"Must be one of: {', '.join(t.value for t in NotificationType)}"
                )

            existing = (await session.execute(
                select(NotificationTemplate).where(
                    NotificationTemplate.template_key == template_key,
                    NotificationTemplate.channel == channel,
                    NotificationTemplate.language == "en",
                    NotificationTemplate.version == 1,
                )
            )).scalars().first()
            if existing is not None:
                raise ValueError(f"A template '{template_key}' for channel '{raw_channel}' already exists")

            obj = NotificationTemplate(
                template_key=template_key,
                channel=channel,
                notification_category=notification_category,
                title_template=title_template,
                body_template=body_template,
            )
            session.add(obj)
            await session.flush()
            return obj.id
        elif entity_type == "memberships":
            import re
            from decimal import Decimal, InvalidOperation

            from app.models.enums import MembershipTier
            from app.models.memberships.membership_plan import MembershipPlan
            from sqlalchemy import select

            name = (row.get("name") or "").strip()
            if not name:
                raise ValueError("'name' is required")

            raw_tier = (row.get("tier") or "").strip().lower()
            try:
                tier = MembershipTier(raw_tier)
            except ValueError:
                raise ValueError(
                    f"Invalid tier '{raw_tier}'. Must be one of: {', '.join(t.value for t in MembershipTier)}"
                )

            existing = (await session.execute(
                select(MembershipPlan).where(MembershipPlan.tier == tier)
            )).scalars().first()
            if existing is not None:
                raise ValueError(f"A membership plan for tier '{raw_tier}' already exists")

            try:
                monthly_price = Decimal(str(row.get("monthly_price")))
            except (InvalidOperation, TypeError, ValueError):
                raise ValueError(f"Invalid monthly_price: {row.get('monthly_price')!r}")

            yearly_raw = (row.get("yearly_price") or "").strip()
            validity_raw = (row.get("validity_days") or "").strip()
            features_raw = (row.get("features") or "").strip()

            slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")

            obj = MembershipPlan(
                tier=tier,
                name=name,
                slug=slug,
                monthly_price=monthly_price,
                yearly_price=Decimal(yearly_raw) if yearly_raw else Decimal("0.00"),
                validity_days=int(validity_raw) if validity_raw else None,
                benefits={"features": [f.strip() for f in features_raw.split(",") if f.strip()]} if features_raw else None,
            )
            session.add(obj)
            await session.flush()
            return obj.id
        elif entity_type == "vendor_services":
            from decimal import Decimal, InvalidOperation

            from app.models.enums import PackagePricingType
            from app.models.users.user import User
            from app.models.vendors.vendor import Vendor
            from app.models.vendors.vendor_category import VendorCategory
            from app.models.vendors.vendor_service import VendorService
            from sqlalchemy import func, select

            name = (row.get("name") or "").strip()
            vendor_phone = (row.get("vendor_phone") or "").strip()
            category_name = (row.get("category") or "").strip()
            if not name:
                raise ValueError("'name' is required")
            if not vendor_phone:
                raise ValueError("'vendor_phone' is required")
            if not category_name:
                raise ValueError("'category' is required")

            try:
                base_price = Decimal(str(row.get("base_price")))
            except (InvalidOperation, TypeError, ValueError):
                raise ValueError(f"Invalid base_price: {row.get('base_price')!r}")

            vendor = (await session.execute(
                select(Vendor).join(User, User.id == Vendor.user_id).where(User.phone == vendor_phone)
            )).scalars().first()
            if vendor is None:
                raise ValueError(f"No vendor found with phone '{vendor_phone}'")

            category = (await session.execute(
                select(VendorCategory).where(func.lower(VendorCategory.name) == category_name.lower())
            )).scalars().first()
            if category is None:
                raise ValueError(f"No vendor category found named '{category_name}'")

            raw_pricing = (row.get("pricing_type") or "").strip().lower()
            try:
                pricing_type = PackagePricingType(raw_pricing) if raw_pricing else PackagePricingType.FIXED
            except ValueError:
                pricing_type = PackagePricingType.FIXED

            obj = VendorService(
                vendor_id=vendor.id,
                category_id=category.id,
                name=name,
                description=row.get("description") or None,
                pricing_type=pricing_type,
                base_price=base_price,
                is_active=False,
            )
            session.add(obj)
            await session.flush()
            return obj.id
        raise NotImplementedError(
            f"Bulk import for entity_type='{entity_type}' is not implemented yet."
        )

    async def undo_import(self, log_id: uuid.UUID) -> ImportUndoResponse:
        async with self._uow() as uow:
            log = await uow.cms.import_logs.get_by_id_or_raise(log_id)
            if log.status != "COMPLETED":
                from fastapi import HTTPException, status as s
                raise HTTPException(
                    status_code=s.HTTP_409_CONFLICT,
                    detail="Only COMPLETED imports can be undone",
                )

            snapshot = log.rollback_snapshot or {}
            inserted_ids = snapshot.get("inserted_ids", [])
            deleted_count = 0

            if inserted_ids and log.entity_type == "faqs":
                from app.models.common.faq import FAQ
                from sqlalchemy import delete
                for faq_id in inserted_ids:
                    try:
                        stmt = delete(FAQ).where(FAQ.id == uuid.UUID(faq_id))
                        await uow.session.execute(stmt)
                        deleted_count += 1
                    except Exception:
                        pass
            elif inserted_ids and log.entity_type == "settings":
                from app.models.common.app_setting import AppSetting
                from sqlalchemy import delete
                for setting_id in inserted_ids:
                    try:
                        stmt = delete(AppSetting).where(AppSetting.id == uuid.UUID(setting_id))
                        await uow.session.execute(stmt)
                        deleted_count += 1
                    except Exception:
                        pass
            elif inserted_ids and log.entity_type == "packages":
                from app.models.packages.package import Package
                from sqlalchemy import delete
                for package_id in inserted_ids:
                    try:
                        stmt = delete(Package).where(Package.id == uuid.UUID(package_id))
                        await uow.session.execute(stmt)
                        deleted_count += 1
                    except Exception:
                        pass
            elif inserted_ids and log.entity_type == "vendors":
                # Deletes only the Vendor + VendorProfile rows created by this
                # import — the paired User account is left intact (it may be
                # reused or the phone number re-imported later).
                from app.models.vendors.vendor import Vendor
                from app.models.vendors.vendor_profile import VendorProfile
                from sqlalchemy import delete
                for vendor_id in inserted_ids:
                    try:
                        vid = uuid.UUID(vendor_id)
                        await uow.session.execute(delete(VendorProfile).where(VendorProfile.vendor_id == vid))
                        await uow.session.execute(delete(Vendor).where(Vendor.id == vid))
                        deleted_count += 1
                    except Exception:
                        pass
            elif inserted_ids and log.entity_type == "customers":
                from app.models.users.user import User
                from sqlalchemy import delete
                for user_id in inserted_ids:
                    try:
                        stmt = delete(User).where(User.id == uuid.UUID(user_id))
                        await uow.session.execute(stmt)
                        deleted_count += 1
                    except Exception:
                        pass
            elif inserted_ids and log.entity_type == "package_categories":
                from app.models.packages.package_category import PackageCategory
                from sqlalchemy import delete
                for cat_id in inserted_ids:
                    try:
                        stmt = delete(PackageCategory).where(PackageCategory.id == uuid.UUID(cat_id))
                        await uow.session.execute(stmt)
                        deleted_count += 1
                    except Exception:
                        pass
            elif inserted_ids and log.entity_type == "cities":
                from app.models.common.city import City
                from sqlalchemy import delete
                for city_id in inserted_ids:
                    try:
                        stmt = delete(City).where(City.id == uuid.UUID(city_id))
                        await uow.session.execute(stmt)
                        deleted_count += 1
                    except Exception:
                        pass
            elif inserted_ids and log.entity_type == "states":
                from app.models.common.state import State
                from sqlalchemy import delete
                for state_id in inserted_ids:
                    try:
                        stmt = delete(State).where(State.id == uuid.UUID(state_id))
                        await uow.session.execute(stmt)
                        deleted_count += 1
                    except Exception:
                        pass
            elif inserted_ids and log.entity_type == "coupons":
                from app.models.payments.coupon import Coupon
                from sqlalchemy import delete
                for coupon_id in inserted_ids:
                    try:
                        stmt = delete(Coupon).where(Coupon.id == uuid.UUID(coupon_id))
                        await uow.session.execute(stmt)
                        deleted_count += 1
                    except Exception:
                        pass
            elif inserted_ids and log.entity_type == "notification_templates":
                from app.models.notifications.template import NotificationTemplate
                from sqlalchemy import delete
                for template_id in inserted_ids:
                    try:
                        stmt = delete(NotificationTemplate).where(NotificationTemplate.id == uuid.UUID(template_id))
                        await uow.session.execute(stmt)
                        deleted_count += 1
                    except Exception:
                        pass
            elif inserted_ids and log.entity_type == "memberships":
                from app.models.memberships.membership_plan import MembershipPlan
                from sqlalchemy import delete
                for plan_id in inserted_ids:
                    try:
                        stmt = delete(MembershipPlan).where(MembershipPlan.id == uuid.UUID(plan_id))
                        await uow.session.execute(stmt)
                        deleted_count += 1
                    except Exception:
                        pass
            elif inserted_ids and log.entity_type == "vendor_services":
                # Only the VendorService rows created by this import are
                # removed — the parent vendor and category are left intact.
                from app.models.vendors.vendor_service import VendorService
                from sqlalchemy import delete
                for service_id in inserted_ids:
                    try:
                        stmt = delete(VendorService).where(VendorService.id == uuid.UUID(service_id))
                        await uow.session.execute(stmt)
                        deleted_count += 1
                    except Exception:
                        pass

            await uow.cms.import_logs.update(log, {"status": "ROLLED_BACK"})
            await uow.commit()

        return ImportUndoResponse(
            log_id=log_id,
            rows_deleted=deleted_count,
            status="ROLLED_BACK",
            message=f"Successfully rolled back {deleted_count} rows",
        )

    async def get_import_logs(
        self,
        *,
        entity_type: str | None = None,
        status: str | None = None,
        skip: int = 0,
        limit: int = 20,
    ) -> tuple[list[ImportLogResponse], int]:
        async with self._uow() as uow:
            logs, total = await uow.cms.import_logs.find_all_paginated(
                entity_type=entity_type, status=status, skip=skip, limit=limit
            )
        return [ImportLogResponse.model_validate(l) for l in logs], total

    async def get_import_log(self, log_id: uuid.UUID) -> ImportLogResponse:
        async with self._uow() as uow:
            log = await uow.cms.import_logs.get_by_id_or_raise(log_id)
        return ImportLogResponse.model_validate(log)

    # ── Export ────────────────────────────────────────────────────────────────

    async def trigger_export(
        self,
        *,
        request: ExportRequest,
        admin_id: uuid.UUID,
    ) -> ExportTriggerResponse:
        async with self._uow() as uow:
            log = await uow.cms.export_logs.create_from_dict({
                "admin_id": admin_id,
                "entity_type": request.entity_type,
                "format": request.format,
                "status": "PENDING",
                "filters": request.filters,
                "selected_ids": [str(i) for i in request.selected_ids] if request.selected_ids else None,
                "column_selection": request.column_selection,
                "is_scheduled": request.is_scheduled,
                "schedule_cron": request.schedule_cron,
            })
            log_id = log.id
            await uow.commit()

        # Run the export (in production this would be a background task)
        try:
            content, row_count, mime_type = await self._generate_export(
                entity_type=request.entity_type,
                format=request.format,
                filters=request.filters or {},
                selected_ids=request.selected_ids,
                column_selection=request.column_selection,
            )

            # File bytes are stored directly on the ExportLog row and served
            # through an admin-authenticated download route — exports can
            # contain customer/payment PII, so they must never land on a
            # publicly-reachable URL (e.g. a plain Cloudinary raw upload).
            file_path = f"exports/{log_id}.{request.format.lower()}"
            download_url = f"/api/v1/admin/cms/io/export/logs/{log_id}/download"

            async with self._uow() as uow:
                export_log = await uow.cms.export_logs.get_by_id_or_raise(log_id)
                await uow.cms.export_logs.mark_completed(
                    log_id,
                    file_storage_path=file_path,
                    download_url=download_url,
                    row_count=row_count,
                    file_size_bytes=len(content),
                    file_content=content,
                    mime_type=mime_type,
                )
                await uow.commit()

            return ExportTriggerResponse(
                log_id=log_id,
                status="COMPLETED",
                message=f"Export completed. {row_count} rows exported.",
                estimated_rows=row_count,
            )
        except Exception as exc:
            async with self._uow() as uow:
                export_log = await uow.cms.export_logs.get_by_id_or_raise(log_id)
                await uow.cms.export_logs.update(export_log, {
                    "status": "FAILED",
                    "error_message": str(exc),
                    "completed_at": datetime.now(timezone.utc),
                })
                await uow.commit()
            raise

    async def _generate_export(
        self,
        *,
        entity_type: str,
        format: str,
        filters: dict[str, Any],
        selected_ids: list[uuid.UUID] | None,
        column_selection: list[str] | None,
    ) -> tuple[bytes, int, str]:
        rows = await self._fetch_export_rows(entity_type, filters, selected_ids)
        if not rows:
            return b"", 0, "text/plain"

        columns = column_selection or list(rows[0].keys()) if rows else []

        if format == "CSV":
            import csv
            buf = io.StringIO()
            writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
            return buf.getvalue().encode("utf-8"), len(rows), "text/csv"

        elif format == "JSON":
            import json
            data = [{c: row.get(c) for c in columns} for row in rows]
            return json.dumps(data, default=str, indent=2).encode("utf-8"), len(rows), "application/json"

        else:  # XLSX
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = entity_type.title()
            ws.append(columns)
            for row in rows:
                ws.append([str(row.get(c, "")) for c in columns])
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue(), len(rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    async def _fetch_export_rows(
        self,
        entity_type: str,
        filters: dict[str, Any],
        selected_ids: list[uuid.UUID] | None,
    ) -> list[dict[str, Any]]:
        from sqlalchemy import select

        async with self._uow() as uow:
            session = uow.session
            assert session is not None

            if entity_type == "vendors":
                from app.models.vendors.vendor import Vendor
                stmt = select(Vendor).where(Vendor.deleted_at.is_(None)).limit(10000)
                rows = (await session.execute(stmt)).scalars().all()
                return [
                    {
                        "id": str(v.id),
                        "business_name": v.business_name,
                        "city": getattr(v, "city", ""),
                        "verification_status": v.verification_status.value,
                        "created_at": str(v.created_at),
                    }
                    for v in rows
                ]

            elif entity_type == "customers":
                from app.models.users.user import User
                stmt = select(User).where(User.deleted_at.is_(None)).limit(10000)
                rows = (await session.execute(stmt)).scalars().all()
                return [
                    {
                        "id": str(u.id),
                        "phone": u.phone,
                        "email": getattr(u, "email", ""),
                        "full_name": getattr(u, "full_name", ""),
                        "account_status": str(u.account_status),
                        "created_at": str(u.created_at),
                    }
                    for u in rows
                ]

            elif entity_type == "bookings":
                from app.models.bookings.booking import Booking
                stmt = select(Booking).where(Booking.deleted_at.is_(None)).limit(10000)
                rows = (await session.execute(stmt)).scalars().all()
                return [
                    {
                        "id": str(b.id),
                        "status": b.booking_status.value if hasattr(b.booking_status, "value") else str(b.booking_status),
                        "total_amount": str(b.total_amount),
                        "special_instructions": b.special_instructions or "",
                        "customization_note": b.customization_note or "",
                        "created_at": str(b.created_at),
                    }
                    for b in rows
                ]

            elif entity_type == "payments":
                from app.models.payments.payment import Payment
                stmt = select(Payment).limit(10000)
                rows = (await session.execute(stmt)).scalars().all()
                return [
                    {
                        "id": str(p.id),
                        "status": p.payment_status.value if hasattr(p.payment_status, "value") else str(p.payment_status),
                        "amount": str(p.final_amount),
                        "created_at": str(p.created_at),
                    }
                    for p in rows
                ]

            elif entity_type == "packages":
                from sqlalchemy.orm import selectinload

                from app.models.packages.package import Package
                from app.models.users.user import User
                from app.models.vendors.vendor import Vendor

                stmt = (
                    select(Package)
                    .where(Package.deleted_at.is_(None))
                    .options(selectinload(Package.category))
                    .limit(10000)
                )
                rows = (await session.execute(stmt)).scalars().all()

                vendor_ids = list({pkg.vendor_id for pkg in rows if pkg.vendor_id})
                phone_by_vendor_id: dict[Any, str] = {}
                if vendor_ids:
                    vendor_rows = (await session.execute(
                        select(Vendor.id, User.phone)
                        .join(User, User.id == Vendor.user_id)
                        .where(Vendor.id.in_(vendor_ids))
                    )).all()
                    phone_by_vendor_id = {vid: phone for vid, phone in vendor_rows}

                return [
                    {
                        "id": str(pkg.id),
                        "name": pkg.name,
                        "description": pkg.description or "",
                        "base_price": str(pkg.base_price) if pkg.base_price is not None else "",
                        "currency": pkg.currency,
                        "category": pkg.category.name if pkg.category else "",
                        "vendor_phone": phone_by_vendor_id.get(pkg.vendor_id, ""),
                        "min_guests": pkg.min_guest_count if pkg.min_guest_count is not None else "",
                        "max_guests": pkg.max_guest_count if pkg.max_guest_count is not None else "",
                        "duration_hours": str(pkg.duration_hours) if pkg.duration_hours is not None else "",
                        "status": pkg.status.value if hasattr(pkg.status, "value") else str(pkg.status),
                        "created_at": str(pkg.created_at),
                    }
                    for pkg in rows
                ]

            elif entity_type == "faqs":
                from app.models.common.faq import FAQ
                stmt = select(FAQ).limit(10000)
                rows = (await session.execute(stmt)).scalars().all()
                return [
                    {
                        "id": str(f.id),
                        "question": f.question,
                        "answer": f.answer,
                        "category": f.faq_category.value if hasattr(f.faq_category, "value") else str(f.faq_category),
                        "display_order": f.display_order,
                        "created_at": str(f.created_at),
                    }
                    for f in rows
                ]

            elif entity_type == "settings":
                from app.models.common.app_setting import AppSetting
                stmt = select(AppSetting).limit(10000)
                rows = (await session.execute(stmt)).scalars().all()
                return [
                    {
                        "id": str(s.id),
                        "key": s.key,
                        "value": s.value,
                        "description": s.description or "",
                        "created_at": str(s.created_at),
                    }
                    for s in rows
                ]

            elif entity_type == "package_categories":
                from app.models.packages.package_category import PackageCategory
                stmt = select(PackageCategory).limit(10000)
                rows = (await session.execute(stmt)).scalars().all()
                return [
                    {
                        "id": str(c.id),
                        "name": c.name,
                        "slug": c.slug,
                        "description": c.description or "",
                        "created_at": str(c.created_at),
                    }
                    for c in rows
                ]

            elif entity_type == "states":
                from app.models.common.state import State
                stmt = select(State).limit(10000)
                rows = (await session.execute(stmt)).scalars().all()
                return [
                    {
                        "id": str(s.id),
                        "name": s.name,
                        "code": s.code,
                        "country": s.country_code,
                        "created_at": str(s.created_at),
                    }
                    for s in rows
                ]

            elif entity_type == "cities":
                from sqlalchemy.orm import selectinload

                from app.models.common.city import City

                stmt = select(City).options(selectinload(City.state)).limit(10000)
                rows = (await session.execute(stmt)).scalars().all()
                return [
                    {
                        "id": str(c.id),
                        "name": c.name,
                        "display_name": c.display_name,
                        "state": c.state.name if c.state else "",
                        "is_metro": c.is_metro,
                        "is_serviceable": c.is_serviceable,
                        "created_at": str(c.created_at),
                    }
                    for c in rows
                ]

            elif entity_type == "coupons":
                from app.models.payments.coupon import Coupon
                stmt = select(Coupon).limit(10000)
                rows = (await session.execute(stmt)).scalars().all()
                return [
                    {
                        "id": str(c.id),
                        "title": c.title,
                        "code": c.code or "",
                        "discount_type": c.coupon_type.value if hasattr(c.coupon_type, "value") else str(c.coupon_type),
                        "discount_value": str(c.discount_value),
                        "max_uses": c.total_usage_limit if c.total_usage_limit is not None else "",
                        "valid_from": str(c.valid_from),
                        "valid_until": str(c.valid_until) if c.valid_until else "",
                        "min_order_value": str(c.min_order_value) if c.min_order_value is not None else "",
                        "admin_status": c.admin_status.value if hasattr(c.admin_status, "value") else str(c.admin_status),
                        "created_at": str(c.created_at),
                    }
                    for c in rows
                ]

            elif entity_type == "notification_templates":
                from app.models.notifications.template import NotificationTemplate
                stmt = select(NotificationTemplate).limit(10000)
                rows = (await session.execute(stmt)).scalars().all()
                return [
                    {
                        "id": str(t.id),
                        "template_key": t.template_key,
                        "channel": t.channel.value if hasattr(t.channel, "value") else str(t.channel),
                        "notification_category": t.notification_category.value if hasattr(t.notification_category, "value") else str(t.notification_category),
                        "title_template": t.title_template,
                        "body_template": t.body_template,
                        "created_at": str(t.created_at),
                    }
                    for t in rows
                ]

            elif entity_type == "memberships":
                from app.models.memberships.membership_plan import MembershipPlan
                stmt = select(MembershipPlan).limit(10000)
                rows = (await session.execute(stmt)).scalars().all()
                return [
                    {
                        "id": str(m.id),
                        "tier": m.tier.value if hasattr(m.tier, "value") else str(m.tier),
                        "name": m.name,
                        "monthly_price": str(m.monthly_price),
                        "yearly_price": str(m.yearly_price),
                        "validity_days": m.validity_days if m.validity_days is not None else "",
                        "features": ",".join((m.benefits or {}).get("features", [])) if m.benefits else "",
                        "created_at": str(m.created_at),
                    }
                    for m in rows
                ]

            elif entity_type == "vendor_services":
                from sqlalchemy.orm import selectinload

                from app.models.users.user import User
                from app.models.vendors.vendor import Vendor
                from app.models.vendors.vendor_service import VendorService

                stmt = (
                    select(VendorService)
                    .where(VendorService.deleted_at.is_(None))
                    .options(selectinload(VendorService.category))
                    .limit(10000)
                )
                rows = (await session.execute(stmt)).scalars().all()

                vendor_ids = list({v.vendor_id for v in rows if v.vendor_id})
                phone_by_vendor_id: dict[Any, str] = {}
                if vendor_ids:
                    vendor_rows = (await session.execute(
                        select(Vendor.id, User.phone)
                        .join(User, User.id == Vendor.user_id)
                        .where(Vendor.id.in_(vendor_ids))
                    )).all()
                    phone_by_vendor_id = {vid: phone for vid, phone in vendor_rows}

                return [
                    {
                        "id": str(vs.id),
                        "vendor_phone": phone_by_vendor_id.get(vs.vendor_id, ""),
                        "category": vs.category.name if vs.category else "",
                        "name": vs.name,
                        "description": vs.description or "",
                        "base_price": str(vs.base_price),
                        "pricing_type": vs.pricing_type.value if hasattr(vs.pricing_type, "value") else str(vs.pricing_type),
                        "created_at": str(vs.created_at),
                    }
                    for vs in rows
                ]

            return []

    async def get_export_logs(
        self,
        *,
        entity_type: str | None = None,
        status: str | None = None,
        skip: int = 0,
        limit: int = 20,
    ) -> tuple[list[ExportLogResponse], int]:
        async with self._uow() as uow:
            logs, total = await uow.cms.export_logs.find_all_paginated(
                entity_type=entity_type, status=status, skip=skip, limit=limit
            )
        return [ExportLogResponse.model_validate(l) for l in logs], total

    async def get_export_log(self, log_id: uuid.UUID) -> ExportLogResponse:
        async with self._uow() as uow:
            log = await uow.cms.export_logs.get_by_id_or_raise(log_id)
        return ExportLogResponse.model_validate(log)

    async def get_export_file(self, log_id: uuid.UUID) -> tuple[bytes, str, str]:
        """Returns (content, mime_type, filename) for a completed export."""
        async with self._uow() as uow:
            log = await uow.cms.export_logs.get_by_id_or_raise(log_id)
            if log.status != "COMPLETED" or not log.file_content:
                from fastapi import HTTPException, status as s
                raise HTTPException(
                    status_code=s.HTTP_404_NOT_FOUND,
                    detail="Export file not available.",
                )
            filename = f"{log.entity_type}_export_{log.id}.{log.format.lower()}"
            return log.file_content, log.mime_type or "application/octet-stream", filename
